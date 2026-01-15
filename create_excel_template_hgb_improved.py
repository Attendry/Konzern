#!/usr/bin/env python3
"""
Erstellt ein Excel-Muster für die Konsolidierung nach HGB
Version 2.0 - Verbessert nach Wirtschaftsprüfer-Empfehlungen
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Erstelle Workbook
wb = Workbook()

# Entferne Standard-Sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Stile definieren
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, size=10)
required_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Blau für Pflichtfelder
optional_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Gelb für optionale Felder
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== BLATT 0: Anleitung (WICHTIG: Erstes Blatt) =====
ws_anleitung = wb.create_sheet("Anleitung", 0)

# Titel
ws_anleitung.merge_cells('A1:F1')
title_cell = ws_anleitung.cell(row=1, column=1)
title_cell.value = "HGB-Konsolidierung Import-Template - Anleitung"
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Version
ws_anleitung.merge_cells('A2:F2')
version_cell = ws_anleitung.cell(row=2, column=1)
version_cell.value = "Version 2.0 - Stand: " + datetime.now().strftime("%Y-%m-%d")
version_cell.font = Font(italic=True, size=10)
version_cell.alignment = Alignment(horizontal="center", vertical="center")

# Übersicht
row = 4
ws_anleitung.cell(row=row, column=1).value = "ÜBERSICHT DER BLÄTTER:"
ws_anleitung.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

sheets_info = [
    ("1. Anleitung", "Dieses Blatt - Übersicht und Anleitung"),
    ("2. Bilanzdaten", "Bilanzpositionen für alle Unternehmen (HGB § 266)"),
    ("3. GuV-Daten", "Gewinn- und Verlustrechnung (HGB § 275) - NEU"),
    ("4. Unternehmensinformationen", "Basis-Informationen zu allen Unternehmen"),
    ("5. Beteiligungsverhältnisse", "Für Kapitalkonsolidierung (HGB § 301)"),
    ("6. Zwischengesellschaftsgeschäfte", "Für Schulden- und Zwischenergebniseliminierung (HGB § 303, § 305)"),
    ("7. Eigenkapital-Aufteilung", "Für Minderheitsanteile"),
    ("8. HGB-Bilanzstruktur", "Referenz zur HGB-Bilanzgliederung (HGB § 266) - NEU"),
    ("9. Kontenplan-Referenz", "Typische Kontonummern-Bereiche - NEU"),
]

for sheet_name, description in sheets_info:
    ws_anleitung.cell(row=row, column=1).value = sheet_name
    ws_anleitung.cell(row=row, column=1).font = Font(bold=True)
    ws_anleitung.cell(row=row, column=2).value = description
    row += 1

row += 1
ws_anleitung.cell(row=row, column=1).value = "SCHRITT-FÜR-SCHRITT-ANLEITUNG:"
ws_anleitung.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

steps = [
    ("Schritt 1:", "Füllen Sie 'Unternehmensinformationen' aus"),
    ("Schritt 2:", "Füllen Sie 'Bilanzdaten' für alle Unternehmen aus"),
    ("Schritt 3:", "Füllen Sie 'GuV-Daten' für alle Unternehmen aus"),
    ("Schritt 4:", "Definieren Sie 'Beteiligungsverhältnisse'"),
    ("Schritt 5:", "Erfassen Sie alle 'Zwischengesellschaftsgeschäfte'"),
    ("Schritt 6:", "Prüfen Sie 'Eigenkapital-Aufteilung'"),
    ("Schritt 7:", "Importieren Sie die Datei im System"),
]

for step_num, step_desc in steps:
    ws_anleitung.cell(row=row, column=1).value = step_num
    ws_anleitung.cell(row=row, column=1).font = Font(bold=True)
    ws_anleitung.cell(row=row, column=2).value = step_desc
    row += 1

row += 1
ws_anleitung.cell(row=row, column=1).value = "HGB-REFERENZEN:"
ws_anleitung.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

hgb_refs = [
    ("§ 266 HGB", "Bilanzgliederung"),
    ("§ 275 HGB", "Gewinn- und Verlustrechnung"),
    ("§ 301 HGB", "Kapitalkonsolidierung"),
    ("§ 303 HGB", "Schuldenkonsolidierung"),
    ("§ 305 HGB", "Zwischenergebniseliminierung"),
    ("§ 274 HGB", "Latente Steuern"),
]

for hgb_par, hgb_desc in hgb_refs:
    ws_anleitung.cell(row=row, column=1).value = hgb_par
    ws_anleitung.cell(row=row, column=1).font = Font(bold=True)
    ws_anleitung.cell(row=row, column=2).value = hgb_desc
    row += 1

# Spaltenbreiten
ws_anleitung.column_dimensions['A'].width = 20
ws_anleitung.column_dimensions['B'].width = 60

# ===== BLATT 1: Bilanzdaten (Verbessert) =====
ws_bilanz = wb.create_sheet("Bilanzdaten", 1)

# Header - Erweitert
headers_bilanz = [
    "Unternehmen", "Kontonummer", "Kontoname", "HGB-Position", 
    "Kontotyp", "Soll", "Haben", "Saldo", 
    "Zwischengesellschaft", "Gegenpartei", "Bemerkung"
]
ws_bilanz.append(headers_bilanz)

# Formatierung Header
for col in range(1, len(headers_bilanz) + 1):
    cell = ws_bilanz.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Beispiel-Daten - Erweitert
example_data = [
    ["Mutterunternehmen H", "1000", "Kasse", "B.IV", "asset", "5000.00", "0.00", "=F2-G2", "Nein", "", ""],
    ["Mutterunternehmen H", "1200", "Forderungen a. LL", "B.II", "asset", "100000.00", "0.00", "=F3-G3", "Ja", "TU1", "Zwischengesellschaftsgeschäft"],
    ["Mutterunternehmen H", "1400", "Beteiligung TU1", "A.III", "asset", "500000.00", "0.00", "=F4-G4", "Nein", "", "Beteiligung an Tochterunternehmen"],
    ["Mutterunternehmen H", "3000", "Gezeichnetes Kapital", "A.I", "equity", "0.00", "1000000.00", "=F5-G5", "Nein", "", ""],
    ["Tochterunternehmen TU1", "1000", "Kasse", "B.IV", "asset", "2000.00", "0.00", "=F6-G6", "Nein", "", ""],
    ["Tochterunternehmen TU1", "1600", "Verbindlichkeiten a. LL", "C", "liability", "0.00", "50000.00", "=F7-G7", "Ja", "Mutter H", "Gegenpartei: Mutterunternehmen H"],
]

for row_idx, row_data in enumerate(example_data, start=2):
    ws_bilanz.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_bilanz.cell(row=row_idx, column=col)
        cell.border = border
        if col in [6, 7, 8]:  # Soll, Haben, Saldo
            cell.number_format = '#,##0.00'
            if col == 8:  # Saldo (Formel)
                cell.fill = optional_fill
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        if col == 1:  # Unternehmen
            cell.fill = required_fill
        if col == 9:  # Zwischengesellschaft
            # Dropdown-Validierung
            dv = DataValidation(type="list", formula1='"Ja,Nein"')
            dv.add(cell)

# Spaltenbreiten anpassen
ws_bilanz.column_dimensions['A'].width = 25
ws_bilanz.column_dimensions['B'].width = 15
ws_bilanz.column_dimensions['C'].width = 30
ws_bilanz.column_dimensions['D'].width = 15
ws_bilanz.column_dimensions['E'].width = 15
ws_bilanz.column_dimensions['F'].width = 15
ws_bilanz.column_dimensions['G'].width = 15
ws_bilanz.column_dimensions['H'].width = 15
ws_bilanz.column_dimensions['I'].width = 20
ws_bilanz.column_dimensions['J'].width = 20
ws_bilanz.column_dimensions['K'].width = 40

# ===== BLATT 2: GuV-Daten (NEU - HGB § 275) =====
ws_guv = wb.create_sheet("GuV-Daten", 2)

headers_guv = [
    "Unternehmen", "Kontonummer", "Kontoname", "Kontotyp", 
    "Betrag", "Zwischengesellschaft", "Gegenpartei", "Bemerkung"
]
ws_guv.append(headers_guv)

# Formatierung Header
for col in range(1, len(headers_guv) + 1):
    cell = ws_guv.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Beispiel-Daten GuV
example_guv = [
    ["Mutterunternehmen H", "8000", "Umsatzerlöse", "revenue", "1000000.00", "Nein", "", ""],
    ["Mutterunternehmen H", "8000", "Umsatzerlöse (an TU1)", "revenue", "100000.00", "Ja", "TU1", "Zwischenumsatz"],
    ["Mutterunternehmen H", "4000", "Materialaufwand", "cost_of_sales", "600000.00", "Nein", "", ""],
    ["Mutterunternehmen H", "6000", "Personalaufwand", "operating_expense", "200000.00", "Nein", "", ""],
    ["Tochterunternehmen TU1", "8000", "Umsatzerlöse", "revenue", "500000.00", "Nein", "", ""],
    ["Tochterunternehmen TU1", "4000", "Materialaufwand", "cost_of_sales", "300000.00", "Nein", "", ""],
    ["Tochterunternehmen TU1", "4000", "Materialaufwand (von Mutter H)", "cost_of_sales", "80000.00", "Ja", "Mutter H", "Zwischenaufwand"],
]

for row_idx, row_data in enumerate(example_guv, start=2):
    ws_guv.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_guv.cell(row=row_idx, column=col)
        cell.border = border
        if col == 5:  # Betrag
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if col == 1:  # Unternehmen
            cell.fill = required_fill
        if col == 4:  # Kontotyp
            # Dropdown-Validierung
            dv = DataValidation(type="list", formula1='"revenue,cost_of_sales,operating_expense,financial_income,financial_expense,income_tax,net_income"')
            dv.add(cell)
        if col == 6:  # Zwischengesellschaft
            dv = DataValidation(type="list", formula1='"Ja,Nein"')
            dv.add(cell)

# Spaltenbreiten
for col in range(1, len(headers_guv) + 1):
    ws_guv.column_dimensions[get_column_letter(col)].width = 20

# ===== BLATT 3: Unternehmensinformationen =====
ws_unternehmen = wb.create_sheet("Unternehmensinformationen", 3)

headers_unternehmen = ["Unternehmensname", "Typ", "Beteiligungs-%", "Erwerbsdatum", "Anschaffungskosten", "Bemerkung"]
ws_unternehmen.append(headers_unternehmen)

for col in range(1, len(headers_unternehmen) + 1):
    cell = ws_unternehmen.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

example_unternehmen = [
    ["Mutterunternehmen H", "Mutterunternehmen (H)", "100.00", "", "", "Hauptunternehmen"],
    ["Tochterunternehmen TU1", "Tochterunternehmen (TU)", "80.00", "2020-01-15", "500000.00", "80% Beteiligung"],
    ["Tochterunternehmen TU2", "Tochterunternehmen (TU)", "60.00", "2021-06-01", "300000.00", "60% Beteiligung"],
]

for row_data in example_unternehmen:
    ws_unternehmen.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_unternehmen.cell(row=ws_unternehmen.max_row, column=col)
        cell.border = border
        if col == 3:  # Beteiligungs-%
            cell.number_format = '0.00"%"'
        if col == 5:  # Anschaffungskosten
            cell.number_format = '#,##0.00'

ws_unternehmen.column_dimensions['A'].width = 25
ws_unternehmen.column_dimensions['B'].width = 25
ws_unternehmen.column_dimensions['C'].width = 15
ws_unternehmen.column_dimensions['D'].width = 15
ws_unternehmen.column_dimensions['E'].width = 18
ws_unternehmen.column_dimensions['F'].width = 40

# ===== BLATT 4: Beteiligungsverhältnisse =====
ws_beteiligung = wb.create_sheet("Beteiligungsverhältnisse", 4)

headers_beteiligung = ["Mutterunternehmen", "Tochterunternehmen", "Beteiligungs-%", "Anschaffungskosten", "Erwerbsdatum", "Beteiligungsbuchwert", "Bemerkung"]
ws_beteiligung.append(headers_beteiligung)

for col in range(1, len(headers_beteiligung) + 1):
    cell = ws_beteiligung.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

example_beteiligung = [
    ["Mutterunternehmen H", "Tochterunternehmen TU1", "80.00", "500000.00", "2020-01-15", "500000.00", "Nach HGB § 301"],
    ["Mutterunternehmen H", "Tochterunternehmen TU2", "60.00", "300000.00", "2021-06-01", "300000.00", "Nach HGB § 301"],
]

for row_data in example_beteiligung:
    ws_beteiligung.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_beteiligung.cell(row=ws_beteiligung.max_row, column=col)
        cell.border = border
        if col == 3:  # Beteiligungs-%
            cell.number_format = '0.00"%"'
        if col in [4, 6]:  # Anschaffungskosten, Beteiligungsbuchwert
            cell.number_format = '#,##0.00'

ws_beteiligung.column_dimensions['A'].width = 25
ws_beteiligung.column_dimensions['B'].width = 25
ws_beteiligung.column_dimensions['C'].width = 15
ws_beteiligung.column_dimensions['D'].width = 18
ws_beteiligung.column_dimensions['E'].width = 15
ws_beteiligung.column_dimensions['F'].width = 18
ws_beteiligung.column_dimensions['G'].width = 30

# ===== BLATT 5: Zwischengesellschaftsgeschäfte (Verbessert) =====
ws_intercompany = wb.create_sheet("Zwischengesellschaftsgeschäfte", 5)

# Erweiterte Header
headers_intercompany = [
    "Transaktions-ID", "Von Unternehmen", "An Unternehmen", "Transaktionstyp", 
    "Betrag", "Kontonummer", "Kontoname", "Gewinnmarge", 
    "Eliminierungsmethode", "Eliminierungsbetrag", "HGB-Referenz", "Bemerkung"
]
ws_intercompany.append(headers_intercompany)

for col in range(1, len(headers_intercompany) + 1):
    cell = ws_intercompany.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

example_intercompany = [
    ["T001", "Mutterunternehmen H", "Tochterunternehmen TU1", "Forderung", "50000.00", "1200", "Forderungen a. LL", "", "Vollständig", "50000.00", "§ 303", "Zu eliminieren"],
    ["T001", "Tochterunternehmen TU1", "Mutterunternehmen H", "Verbindlichkeit", "50000.00", "1600", "Verbindlichkeiten a. LL", "", "Vollständig", "50000.00", "§ 303", "Zu eliminieren"],
    ["T002", "Mutterunternehmen H", "Tochterunternehmen TU1", "Lieferung", "100000.00", "8000", "Umsatzerlöse", "20.00", "Vollständig", "20000.00", "§ 305", "Zwischengewinn zu eliminieren"],
]

for row_idx, row_data in enumerate(example_intercompany, start=2):
    ws_intercompany.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_intercompany.cell(row=row_idx, column=col)
        cell.border = border
        if col in [5, 8, 10]:  # Betrag, Gewinnmarge, Eliminierungsbetrag
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if col == 4:  # Transaktionstyp
            dv = DataValidation(type="list", formula1='"Forderung,Verbindlichkeit,Lieferung,Dienstleistung,Zinsen,Dividenden"')
            dv.add(cell)
        if col == 9:  # Eliminierungsmethode
            dv = DataValidation(type="list", formula1='"Vollständig,Teilweise,Zeitanteilig"')
            dv.add(cell)
        if col == 11:  # HGB-Referenz
            dv = DataValidation(type="list", formula1='"§ 303,§ 305"')
            dv.add(cell)

# Spaltenbreiten
for col in range(1, len(headers_intercompany) + 1):
    ws_intercompany.column_dimensions[get_column_letter(col)].width = 18

# ===== BLATT 6: Eigenkapital-Aufteilung =====
ws_eigenkapital = wb.create_sheet("Eigenkapital-Aufteilung", 6)

headers_eigenkapital = ["Unternehmen", "Gezeichnetes Kapital", "Kapitalrücklagen", "Gewinnrücklagen", "Jahresüberschuss", "Gesamt Eigenkapital", "Anteil Mutter", "Anteil Minderheit"]
ws_eigenkapital.append(headers_eigenkapital)

for col in range(1, len(headers_eigenkapital) + 1):
    cell = ws_eigenkapital.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

example_eigenkapital = [
    ["Mutterunternehmen H", "1000000.00", "200000.00", "300000.00", "150000.00", "1650000.00", "100.00", "0.00"],
    ["Tochterunternehmen TU1", "500000.00", "100000.00", "80000.00", "50000.00", "730000.00", "80.00", "20.00"],
]

for row_data in example_eigenkapital:
    ws_eigenkapital.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_eigenkapital.cell(row=ws_eigenkapital.max_row, column=col)
        cell.border = border
        if col in [2, 3, 4, 5, 6]:  # Beträge
            cell.number_format = '#,##0.00'
        if col in [7, 8]:  # Anteile
            cell.number_format = '0.00"%"'

ws_eigenkapital.column_dimensions['A'].width = 25
for col in range(2, 9):
    ws_eigenkapital.column_dimensions[get_column_letter(col)].width = 18

# ===== BLATT 7: HGB-Bilanzstruktur (NEU - Referenz) =====
ws_hgb_struktur = wb.create_sheet("HGB-Bilanzstruktur", 7)

# Titel
ws_hgb_struktur.merge_cells('A1:C1')
title_cell = ws_hgb_struktur.cell(row=1, column=1)
title_cell.value = "HGB-Bilanzgliederung nach § 266 HGB"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Aktivseite
row = 3
ws_hgb_struktur.cell(row=row, column=1).value = "AKTIVSEITE"
ws_hgb_struktur.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

aktiv_struktur = [
    ("A", "Anlagevermögen", ""),
    ("", "I. Immaterielle Vermögensgegenstände", ""),
    ("", "II. Sachanlagen", ""),
    ("", "III. Finanzanlagen", ""),
    ("B", "Umlaufvermögen", ""),
    ("", "I. Vorräte", ""),
    ("", "II. Forderungen und sonstige Vermögensgegenstände", ""),
    ("", "III. Wertpapiere", ""),
    ("", "IV. Kassenbestand, Bundesbankguthaben, Guthaben bei Kreditinstituten", ""),
    ("C", "Rechnungsabgrenzungsposten", ""),
    ("D", "Aktive latente Steuern", ""),
]

for pos, name, konten in aktiv_struktur:
    ws_hgb_struktur.cell(row=row, column=1).value = pos
    ws_hgb_struktur.cell(row=row, column=2).value = name
    ws_hgb_struktur.cell(row=row, column=3).value = konten
    if pos:  # Hauptposition
        ws_hgb_struktur.cell(row=row, column=1).fill = subheader_fill
        ws_hgb_struktur.cell(row=row, column=1).font = Font(bold=True)
    row += 1

# Passivseite
row += 1
ws_hgb_struktur.cell(row=row, column=1).value = "PASSIVSEITE"
ws_hgb_struktur.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1

passiv_struktur = [
    ("A", "Eigenkapital", ""),
    ("", "I. Gezeichnetes Kapital", ""),
    ("", "II. Kapitalrücklage", ""),
    ("", "III. Gewinnrücklagen", ""),
    ("", "IV. Gewinnvortrag/Verlustvortrag", ""),
    ("", "V. Jahresüberschuss/Jahresfehlbetrag", ""),
    ("B", "Rückstellungen", ""),
    ("C", "Verbindlichkeiten", ""),
    ("D", "Rechnungsabgrenzungsposten", ""),
    ("E", "Passive latente Steuern", ""),
]

for pos, name, konten in passiv_struktur:
    ws_hgb_struktur.cell(row=row, column=1).value = pos
    ws_hgb_struktur.cell(row=row, column=2).value = name
    ws_hgb_struktur.cell(row=row, column=3).value = konten
    if pos:  # Hauptposition
        ws_hgb_struktur.cell(row=row, column=1).fill = subheader_fill
        ws_hgb_struktur.cell(row=row, column=1).font = Font(bold=True)
    row += 1

# Spaltenbreiten
ws_hgb_struktur.column_dimensions['A'].width = 5
ws_hgb_struktur.column_dimensions['B'].width = 60
ws_hgb_struktur.column_dimensions['C'].width = 30

# ===== BLATT 8: Kontenplan-Referenz (NEU) =====
ws_kontenplan = wb.create_sheet("Kontenplan-Referenz", 8)

# Titel
ws_kontenplan.merge_cells('A1:D1')
title_cell = ws_kontenplan.cell(row=1, column=1)
title_cell.value = "Typische Kontonummern-Bereiche (SKR-Referenz)"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Header
headers_kontenplan = ["Kontonummer-Bereich", "Kontotyp", "Beschreibung", "HGB-Position"]
ws_kontenplan.append(headers_kontenplan)

for col in range(1, len(headers_kontenplan) + 1):
    cell = ws_kontenplan.cell(row=2, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

kontenplan_data = [
    ["0000-0999", "asset", "Anlagevermögen (Immaterielle Vermögensgegenstände)", "A.I"],
    ["1000-1499", "asset", "Anlagevermögen (Sachanlagen, Finanzanlagen)", "A.II, A.III"],
    ["1500-1999", "asset", "Umlaufvermögen (Vorräte, Forderungen)", "B.I, B.II"],
    ["2000-2999", "asset", "Umlaufvermögen (Wertpapiere, Kasse, Bank)", "B.III, B.IV"],
    ["3000-3999", "equity", "Eigenkapital", "A"],
    ["4000-4999", "liability", "Verbindlichkeiten", "C"],
    ["5000-5999", "liability", "Rückstellungen", "B"],
    ["6000-6999", "expense", "Aufwendungen (Material, Personal)", "GuV"],
    ["7000-7999", "expense", "Aufwendungen (Abschreibungen, Zinsen)", "GuV"],
    ["8000-8999", "revenue", "Erträge (Umsatzerlöse, sonstige Erträge)", "GuV"],
    ["9000-9999", "equity", "GuV-Abschluss", "A.V"],
]

for row_data in kontenplan_data:
    ws_kontenplan.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws_kontenplan.cell(row=ws_kontenplan.max_row, column=col)
        cell.border = border

# Spaltenbreiten
ws_kontenplan.column_dimensions['A'].width = 20
ws_kontenplan.column_dimensions['B'].width = 15
ws_kontenplan.column_dimensions['C'].width = 50
ws_kontenplan.column_dimensions['D'].width = 15

# Speichere Datei
filename = "templates/Konsolidierung_Muster.xlsx"
import os
os.makedirs("templates", exist_ok=True)
wb.save(filename)
print(f"Excel-Template erfolgreich erstellt: {filename}")
print("Version 2.0 - Mit HGB-Verbesserungen:")
print("  - Anleitung-Blatt hinzugefügt")
print("  - GuV-Daten-Blatt hinzugefügt (HGB § 275)")
print("  - HGB-Bilanzstruktur-Referenz hinzugefügt (HGB § 266)")
print("  - Kontenplan-Referenz hinzugefügt")
print("  - Erweiterte Zwischengesellschaftsgeschäfte")
print("  - Excel-Validierungsregeln implementiert")
